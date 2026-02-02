# ci/summarize_xlsx_openai.py
from __future__ import annotations

import argparse
import json
import os
import random
import shutil
import ssl
import sys
import time
import urllib.error
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl

OPENAI_API_URL = "https://api.openai.com/v1/responses"


def _log(msg: str) -> None:
    print(msg, flush=True)


def extract_response_text(data: dict) -> str:
    if isinstance(data.get("output_text"), str) and data["output_text"].strip():
        return data["output_text"].strip()

    for item in data.get("output", []) or []:
        for c in item.get("content", []) or []:
            t = c.get("text")
            if isinstance(t, str) and t.strip():
                return t.strip()
            if c.get("type") == "output_text":
                t2 = c.get("text")
                if isinstance(t2, str) and t2.strip():
                    return t2.strip()
            if c.get("type") == "refusal":
                r = c.get("refusal")
                if isinstance(r, str) and r.strip():
                    return r.strip()
    return ""


def call_openai(
    prompt: str,
    model: str,
    timeout_s: int = 45,
    max_output_tokens: int = 900,
    reasoning_effort: str = "low",  # low|medium|high
) -> str:
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        raise RuntimeError("Missing OPENAI_API_KEY env var")

    body: Dict[str, object] = {
        "model": model,
        "input": prompt,
        "max_output_tokens": max_output_tokens,
        "truncation": "auto",
        # Crucial: keep reasoning from consuming the entire output budget.
        "reasoning": {"effort": reasoning_effort},
        # Force plain text output (and avoid tool-call outputs)
        "text": {"format": {"type": "text"}},
        "tool_choice": "none",
        "parallel_tool_calls": False,
    }

    req = urllib.request.Request(
        OPENAI_API_URL,
        data=json.dumps(body).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )

    ctx = ssl._create_unverified_context()
    with urllib.request.urlopen(req, timeout=timeout_s, context=ctx) as resp:
        data = json.loads(resp.read().decode("utf-8"))

    text = extract_response_text(data)

    # If still empty, surface the real reason immediately.
    if not text.strip():
        preview = json.dumps(data, ensure_ascii=False)[:2500]
        raise RuntimeError(f"Empty model output. Response preview: {preview}")

    return text.strip()


def call_openai_with_retries(
    prompt: str,
    model: str,
    timeout_s: int = 45,
    max_output_tokens: int = 900,
    reasoning_effort: str = "low",
    retries: int = 6,
) -> str:
    backoff = 1.0
    last_body = ""

    for attempt in range(1, retries + 1):
        try:
            return call_openai(
                prompt=prompt,
                model=model,
                timeout_s=timeout_s,
                max_output_tokens=max_output_tokens,
                reasoning_effort=reasoning_effort,
            )

        except urllib.error.HTTPError as e:
            body = ""
            try:
                body = e.read().decode("utf-8", errors="replace")
            except Exception:
                body = "<could not read error body>"
            last_body = body

            if e.code in (400, 401, 403, 404, 422):
                _log(f"HTTP {e.code} (non-retryable). Body:\n{body}")
                raise

            if e.code in (429, 500, 502, 503, 504):
                sleep_s = backoff + random.uniform(0, 0.25)
                _log(
                    f"HTTP {e.code} retry (attempt {attempt}/{retries}) sleeping {sleep_s:.2f}s"
                )
                time.sleep(sleep_s)
                backoff = min(backoff * 1.8, 20)
                continue

            _log(f"HTTP {e.code} unexpected. Body:\n{body}")
            raise

        except (urllib.error.URLError, TimeoutError) as e:
            sleep_s = backoff + random.uniform(0, 0.25)
            _log(
                f"network/timeout (attempt {attempt}/{retries}) sleeping {sleep_s:.2f}s: {e!r}"
            )
            time.sleep(sleep_s)
            backoff = min(backoff * 1.8, 20)

        except RuntimeError as e:
            _log(f"RuntimeError (attempt {attempt}/{retries}): {e}")
            raise

    raise RuntimeError(
        f"OpenAI request failed after retries. Last error body:\n{last_body}"
    )


def _trim_text(full_text: str, max_chars: int = 12000) -> str:
    t = (full_text or "").strip()
    if len(t) <= max_chars:
        return t
    return t[:8000] + "\n\n[...]\n\n" + t[-2000:]


def summarize_row(
    full_text: str,
    model: str,
    timeout_s: int,
    max_output_tokens: int,
    reasoning_effort: str,
) -> str:
    full_text = _trim_text(full_text)

    prompt = (
        "Summarize this page for competitive intel.\n"
        "FORMAT (exact):\n"
        "<Title Case, <= 10 words>\n"
        "\n"
        "• ...\n"
        "Rules:\n"
        "- Concise headline\n"
        "- 5–8 bullets\n"
        "- concrete capabilities / positioning / proof only\n"
        "- no fluff, no intro, no conclusion\n\n"
        f"PAGE TEXT (markdown):\n{full_text}\n"
    )

    return call_openai_with_retries(
        prompt=prompt,
        model=model,
        timeout_s=timeout_s,
        max_output_tokens=max_output_tokens,
        reasoning_effort=reasoning_effort,
        retries=6,
    ).strip()


def build_jobs(
    ws,
    start_row: int,
    text_col: int,
    out_col: int,
    heartbeat_every: int = 200,
) -> List[Tuple[int, str]]:
    jobs: List[Tuple[int, str]] = []
    max_row = ws.max_row
    _log(
        f"Sheet scan: rows {start_row}..{max_row} | text_col={text_col} out_col={out_col}"
    )

    for r in range(start_row, max_row + 1):
        if heartbeat_every and (r % heartbeat_every == 0):
            _log(f"scanned up to row {r}/{max_row}")

        full_text = ws.cell(r, text_col).value
        existing = ws.cell(r, out_col).value

        if existing is not None and str(existing).strip():
            continue
        if full_text is None or not str(full_text).strip():
            continue

        jobs.append((r, str(full_text)))

    return jobs


def _prepare_output_path(xlsx_in: str, out_arg: str) -> Path:
    in_path = Path(xlsx_in).resolve()
    if out_arg.strip():
        out_path = Path(out_arg).resolve()
        out_path.parent.mkdir(parents=True, exist_ok=True)
        if not out_path.exists():
            shutil.copy2(in_path, out_path)
        return out_path
    return in_path


def _verify_cell(xlsx_path: Path, row: int, col: int) -> int:
    wb2 = openpyxl.load_workbook(str(xlsx_path))
    ws2 = wb2.active
    v = ws2.cell(row, col).value
    return len(str(v or ""))


def summarize_xlsx(
    xlsx_in_path: str,
    out_arg: str,
    model: str,
    workers: int = 8,
    start_row: int = 3,
    text_col: int = 4,
    out_col: int = 5,
    timeout_s: int = 45,
    max_output_tokens: int = 1200,
    reasoning_effort: str = "low",
    save_every: int = 10,
) -> None:
    in_path = Path(xlsx_in_path).resolve()
    out_path = _prepare_output_path(xlsx_in_path, out_arg)

    _log(f"Input:  {in_path}")
    _log(f"Output: {out_path}")
    _log(f"Loading workbook for write: {out_path}")

    wb = openpyxl.load_workbook(str(out_path))
    ws = wb.active

    jobs = build_jobs(ws, start_row=start_row, text_col=text_col, out_col=out_col)
    _log(f"jobs: {len(jobs)}")
    if not jobs:
        _log(
            "No rows to summarize (either output col already filled or text col empty)."
        )
        wb.save(str(out_path))
        return

    first_r, first_txt = jobs[0]
    _log(f"first job row: {first_r} text_len: {len(first_txt)}")
    _log(
        f"Running model={model} workers={workers} timeout_s={timeout_s} max_output_tokens={max_output_tokens} reasoning_effort={reasoning_effort}"
    )

    completed = 0

    def write_row(r: int, summary: str) -> None:
        ws.cell(r, out_col).value = summary

    if workers <= 1:
        for r, txt in jobs:
            _log(f"row {r}: start")
            summary = summarize_row(
                txt,
                model=model,
                timeout_s=timeout_s,
                max_output_tokens=max_output_tokens,
                reasoning_effort=reasoning_effort,
            )
            _log(f"row {r}: summary_len={len(summary)}")
            write_row(r, summary)
            completed += 1

            if save_every and (completed % save_every == 0):
                wb.save(str(out_path))
                vlen = _verify_cell(out_path, r, out_col)
                _log(
                    f"checkpoint saved -> {out_path} | verify row {r} col {out_col} value_len={vlen}"
                )

        wb.save(str(out_path))
        vlen = _verify_cell(out_path, first_r, out_col)
        _log(
            f"FINAL saved -> {out_path} | verify row {first_r} col {out_col} value_len={vlen}"
        )
        return

    max_inflight = max(2, workers * 2)
    idx = 0

    def submit(ex, job):
        r, txt = job
        return ex.submit(
            summarize_row, txt, model, timeout_s, max_output_tokens, reasoning_effort
        ), r

    with ThreadPoolExecutor(max_workers=workers) as ex:
        inflight: List[Tuple[object, int]] = []

        while idx < len(jobs) and len(inflight) < max_inflight:
            fut, r = submit(ex, jobs[idx])
            _log(f"row {r}: start")
            inflight.append((fut, r))
            idx += 1

        while inflight:
            for fut in as_completed([f for f, _ in inflight], timeout=None):
                r = next(rr for ff, rr in inflight if ff is fut)
                inflight = [(ff, rr) for ff, rr in inflight if ff is not fut]

                try:
                    summary = fut.result()
                except Exception as e:
                    _log(f"FAILED row {r}: {e!r}")
                    wb.save(str(out_path))
                    _log(f"Saved partial progress -> {out_path}")
                    raise

                _log(f"row {r}: summary_len={len(summary)}")
                write_row(r, summary)
                completed += 1

                if save_every and (completed % save_every == 0):
                    wb.save(str(out_path))
                    vlen = _verify_cell(out_path, r, out_col)
                    _log(
                        f"checkpoint saved -> {out_path} | verify row {r} col {out_col} value_len={vlen}"
                    )

                if idx < len(jobs):
                    fut2, r2 = submit(ex, jobs[idx])
                    _log(f"row {r2}: start")
                    inflight.append((fut2, r2))
                    idx += 1

                break

    wb.save(str(out_path))
    vlen = _verify_cell(out_path, first_r, out_col)
    _log(
        f"FINAL saved -> {out_path} | completed={completed} | verify row {first_r} col {out_col} value_len={vlen}"
    )


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(line_buffering=True)
    except Exception:
        pass

    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--out", default="")
    ap.add_argument("--model", default="gpt-5-mini")
    ap.add_argument("--workers", type=int, default=8)
    ap.add_argument("--start-row", type=int, default=3)
    ap.add_argument("--text-col", type=int, default=4)
    ap.add_argument("--out-col", type=int, default=5)
    ap.add_argument("--timeout", type=int, default=45)
    ap.add_argument("--max-output-tokens", type=int, default=900)
    ap.add_argument(
        "--reasoning-effort", default="low", choices=["low", "medium", "high"]
    )
    ap.add_argument("--save-every", type=int, default=10)
    args = ap.parse_args()

    summarize_xlsx(
        xlsx_in_path=args.xlsx,
        out_arg=args.out,
        model=args.model,
        workers=args.workers,
        start_row=args.start_row,
        text_col=args.text_col,
        out_col=args.out_col,
        timeout_s=args.timeout,
        max_output_tokens=args.max_output_tokens,
        reasoning_effort=args.reasoning_effort,
        save_every=args.save_every,
    )
