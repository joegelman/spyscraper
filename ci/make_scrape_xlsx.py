# ci/make_scrape_xlsx.py
from __future__ import annotations

import json
from pathlib import Path
from urllib.parse import urlparse

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


HEADERS = ["url_slug", "title", "meta_description", "full_text", "summary", "out_links"]


def _load_pages_jsonl(p: str) -> list[dict]:
    rows = []
    with open(p, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def _url_slug(url: str) -> str:
    try:
        return urlparse(url).path or "/"
    except Exception:
        return ""


def write_scrape_xlsx(
    pages_jsonl: str,
    out_xlsx: str,
    title: str,
    include_all: bool = True,
):
    pages = _load_pages_jsonl(pages_jsonl)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pages"

    # Row 1: title (merged)
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))

    # Row 2: headers
    for i, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Data rows start at row 3
    r = 3
    for p in pages:
        url = p.get("final_url") or p.get("url") or ""
        ws.cell(r, 1, _url_slug(url))
        ws.cell(r, 2, p.get("title") or "")
        ws.cell(r, 3, p.get("meta_description") or "")

        # full_text should be markdown display text, not raw html
        full_text = p.get("text_md") or p.get("text") or ""
        ws.cell(r, 4, full_text)

        # summary column is filled later by OpenAI (leave blank now)
        ws.cell(r, 5, "")

        # store out_links as a JSON string like your sample
        out_links = p.get("links") or []
        ws.cell(r, 6, json.dumps(out_links, ensure_ascii=False))

        r += 1

    # basic formatting
    widths = [22, 38, 48, 90, 55, 45]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A3"
    Path(out_xlsx).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)
    print(f"Wrote {out_xlsx}")


if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser()
    ap.add_argument("--pages", required=True, help="crawl/pages.jsonl")
    ap.add_argument("--out", required=True, help="output .xlsx path")
    ap.add_argument("--title", required=True, help="title row text")
    args = ap.parse_args()

    write_scrape_xlsx(args.pages, args.out, args.title)

